import openpyxl

wb = openpyxl.Workbook()

wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

cell = sheet["a1"]
cell = sheet.cell(row=1, column=1)
for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)
column = sheet["a"]
cells = sheet["a:c"]
sheet["a1:c3"]
sheet[1:3]

wb.create_sheet("Sheet2", 0)
wb.remove_sheet(sheet)